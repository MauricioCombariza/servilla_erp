from __future__ import annotations

import io
import re
import unicodedata
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.personal import Personal
from app.services.bases_web import fetch_pendientes_courier

_MESES_ES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]

COLUMNAS_EXCEL = [
    "serial", "orden", "cod_men", "f_emi", "no_entidad", "nombred",
    "dirdes1", "cod_sec", "ciudad1", "dpto1", "retorno", "ret_esc", "motivo",
]

_NOMBRE_INVALIDO_RE = re.compile(r"[^A-Za-z0-9_-]")


def _slug_archivo(nombre: str) -> str:
    normalizado = unicodedata.normalize("NFKD", nombre).encode("ascii", "ignore").decode("ascii")
    return _NOMBRE_INVALIDO_RE.sub("", normalizado.replace(" ", "_")) or "courier"


async def _get_personal_activo(db: AsyncSession, tipo_personal: str) -> list[Personal]:
    result = await db.execute(
        select(Personal)
        .where(Personal.tipo_personal == tipo_personal, Personal.activo == True)  # noqa: E712
        .order_by(Personal.nombre_completo)
    )
    return list(result.scalars().all())


def _codigos_numericos(personas: list[Personal]) -> dict[int, Personal]:
    """Mapea codigo numérico -> Personal, descartando códigos no numéricos
    (ej. 'CE99'), igual que updateDashboard.py::cargar_courier_externo."""
    out: dict[int, Personal] = {}
    for p in personas:
        try:
            out[int(p.codigo)] = p
        except (TypeError, ValueError):
            continue
    return out


def _agrupar_por_cod_men(rows: list[dict]) -> dict[str, list[dict]]:
    """Agrupa filas de histo por cod_men y deduplica por serial dentro de cada
    grupo (igual a drop_duplicates('serial', keep='first') del script original)."""
    agrupado: dict[str, list[dict]] = defaultdict(list)
    vistos: dict[str, set] = defaultdict(set)
    for row in rows:
        cod_men = str(row["cod_men"])
        serial = row["serial"]
        if serial in vistos[cod_men]:
            continue
        vistos[cod_men].add(serial)
        agrupado[cod_men].append(row)
    return agrupado


def calcular_desglose_mensual(rows: list[dict]) -> list[dict]:
    conteo: dict[str, int] = defaultdict(int)
    for row in rows:
        anomes = (row.get("f_emi") or "")[:7]  # 'YYYY.MM'
        if len(anomes) != 7 or anomes[4] != ".":
            continue
        conteo[anomes] += 1

    desglose = []
    for anomes in sorted(conteo):
        anio, mes = anomes.split(".")
        try:
            nombre_mes = _MESES_ES[int(mes) - 1]
        except (ValueError, IndexError):
            continue
        desglose.append({"mes": f"{nombre_mes} {anio}", "pendientes": conteo[anomes]})
    return desglose


async def get_pendientes_por_personal(
    db: AsyncSession, tipo_personal: str, dias_corte: int = 60,
) -> list[dict]:
    personas = await _get_personal_activo(db, tipo_personal)
    codigos_map = _codigos_numericos(personas)
    if not codigos_map:
        return []

    rows = await fetch_pendientes_courier(list(codigos_map.keys()), dias_corte)
    agrupado = _agrupar_por_cod_men(rows)

    resultado = []
    for codigo, persona in codigos_map.items():
        filas = agrupado.get(str(codigo), [])
        if not filas:
            continue
        email = persona.email
        resultado.append({
            "cod_men": persona.codigo,
            "nombre": persona.nombre_completo,
            "email": email,
            "tiene_email": bool(email and "@" in email),
            "pendientes": len(filas),
            "desglose_mensual": calcular_desglose_mensual(filas),
        })
    resultado.sort(key=lambda r: r["nombre"])
    return resultado


async def get_pendientes_courier_individual(
    db: AsyncSession, codigo: str, tipo_personal: str, dias_corte: int = 60,
) -> tuple[Personal, list[dict]] | None:
    result = await db.execute(
        select(Personal).where(
            Personal.codigo == codigo,
            Personal.tipo_personal == tipo_personal,
            Personal.activo == True,  # noqa: E712
        )
    )
    persona = result.scalar_one_or_none()
    if persona is None:
        return None
    try:
        codigo_int = int(persona.codigo)
    except (TypeError, ValueError):
        return persona, []

    rows = await fetch_pendientes_courier([codigo_int], dias_corte)
    agrupado = _agrupar_por_cod_men(rows)
    return persona, agrupado.get(str(codigo_int), [])


def nombre_archivo_excel(persona: Personal) -> str:
    return f"pendientes_{persona.codigo}_{_slug_archivo(persona.nombre_completo)}.xlsx"


def construir_excel_courier(nombre: str, filas: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pendientes"

    titulo = "Pendientes de entrega" + (f" - {nombre}" if nombre else "")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNAS_EXCEL))
    ws.cell(row=1, column=1, value=titulo).font = Font(bold=True, size=13)

    header_row = 3
    for col, nombre_col in enumerate(COLUMNAS_EXCEL, start=1):
        c = ws.cell(row=header_row, column=col, value=nombre_col)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    row_idx = header_row + 1
    for fila in filas:
        for col, key in enumerate(COLUMNAS_EXCEL, start=1):
            ws.cell(row=row_idx, column=col, value=fila.get(key))
        row_idx += 1

    widths = [16, 10, 10, 12, 26, 26, 30, 10, 16, 12, 10, 10, 20]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
