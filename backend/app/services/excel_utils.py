from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def construir_excel(titulo: str, columnas: list[str], filas: list[dict], widths: list[int]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columnas))
    ws.cell(row=1, column=1, value=titulo).font = Font(bold=True, size=13)

    header_row = 3
    for col, nombre_col in enumerate(columnas, start=1):
        c = ws.cell(row=header_row, column=col, value=nombre_col)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    row_idx = header_row + 1
    for fila in filas:
        for col, key in enumerate(columnas, start=1):
            ws.cell(row=row_idx, column=col, value=fila.get(key))
        row_idx += 1

    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
